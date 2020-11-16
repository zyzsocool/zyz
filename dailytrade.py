import pandas as pd
import os
import datetime
from openpyxl import load_workbook
import numpy as np
pd.set_option('display.max_columns', None)#显示所有列
pd.set_option('display.max_rows', None)#显示所有行
pd.set_option('display.width',1000)
pd.set_option('display.unicode.ambiguous_as_wide', True)
pd.set_option('display.unicode.east_asian_width', True)

while True:
    day = input('输入交易日（如20201113），按回车键结束')
    lastday = input('输入上个交易日（如20201112），按回车键结束')
    confirm=input('是否确认交易日（输入1确认，输入其他则更改日期），按回车键结束')
    if confirm=='1':
        break
#day='20201113'
day_dt=datetime.datetime.strptime(day,'%Y%m%d')
lastday='20201112'
address='E://衍生品日报'
address1 = address + '//' + day
address2 = address + '//' + lastday


outaddress=address1+'//衍生品日报'+day+'.xlsx'

file1 = os.listdir(address1)
file2 = os.listdir(address2)
df_bond =pd.DataFrame()
df_irs =pd.DataFrame()
df_spot =pd.DataFrame()
df_forward =pd.DataFrame()
df_forward = pd.DataFrame()
df_swap = pd.DataFrame()
df_option =pd.DataFrame()
df_borrow =pd.DataFrame()
def mymean(a, b):
    a = list(a)
    b = list(b)
    k = len(a)
    p = 0
    for i in range(k):
        p += a[i] * b[i]
    return p / sum(b)
for i in file1:
    if '现券' in i:
        df_bond = pd.read_excel(address1 + '//' + i, header=1)
    elif 'irs' in i:
        df_irs = pd.read_excel(address1 + '//' + i, header=1)
    elif '即期投组交易查询与维护' in i:
        df_spot = pd.read_excel(address1 + '//' + i, header=1)
    elif '远期投组交易查询与维护' in i:
        df_forward = pd.read_excel(address1 + '//' + i, header=1)
    elif '掉期投组交易查询与维护' in i:
        df_swap = pd.read_excel(address1 + '//' + i, header=1)
    elif '期权投组交易查询与维护' in i:
        df_option = pd.read_excel(address1 + '//' + i, header=1)
    elif '拆借投组交易查询与维护' in i:
        df_borrow = pd.read_excel(address1 + '//' + i, header=1)
    elif '收益风险评估' in i:
        df_dvtoday = pd.read_excel(address1 + '//' + i, header=0)
        df_dvtoday.rename(columns={'Unnamed: 2':'成交编号/债券'},inplace=True)
    elif '年逐日盯市损益分析' in i:
        df_year= pd.read_excel(address1 + '//' + i, header=0)
    elif '季逐日盯市损益分析' in i:
        df_quarter=pd.read_excel(address1 + '//' + i, header=0)
    elif '敞口统计与分析' in i :
        df_fxpositon=pd.read_excel(address1 + '//' + i, header=0)
    elif '损益统计与分析' in i:
        df_fxgain=pd.read_excel(address1 + '//' + i, header=0)
for i in file2:
    if '收益风险评估' in i:
        df_dvlastday = pd.read_excel(address2 + '//' + i, header=0)
        df_dvlastday.rename(columns={'Unnamed: 2':'成交编号/债券'},inplace=True)
if df_irs.empty==False:
    df_irs=pd.merge(df_irs,df_dvtoday,how='left',left_on='成交编号',right_on='成交编号/债券').copy()
    df_irs=df_irs[['投资组合','交易方向','名义本金(万)','期限','到期日_x','本方收取_固定利率(%)','本方支付_固定利率(%)','成交编号','PVBP']].copy()
    df_irs['交易方向']=df_irs['交易方向'].apply(lambda x:'BUY' if x=='收浮动 付固定'  else 'SELL')
    df_irs['期限']=df_irs['期限']/365
    df_irs['PVBP']=df_irs['PVBP']/10000
    df_irs.insert(1,'交易工具','IRS')
    df_irs.insert(5,'利率/汇率/汇差',df_irs.apply(lambda x:x['本方收取_固定利率(%)'] if type(x['本方收取_固定利率(%)'])==float else x['本方支付_固定利率(%)'],axis=1))
    df_irs.insert(6,'近端日/行权日','-')
    df_irs=df_irs.drop(labels=['本方收取_固定利率(%)','本方支付_固定利率(%)','成交编号'],axis=1).copy()
    df_irs['交易类型']=df_irs['期限'].apply(lambda x: 'FR007S5Y' if x>4.8 else 'FR007S1Y')
    df_irs.rename(columns={'名义本金(万)':'面额/名义本金','期限':'剩余期限','到期日_x':'远端日/交割日/到期日'},inplace=True)
    #print(df_irs)

if df_bond.empty==False:
    df_bond_buy=df_bond[df_bond['交易方向']=='买入']
    df_bond_sell=df_bond[df_bond['交易方向']=='卖出']
    df_bond_buy=pd.merge(df_bond_buy,df_dvtoday,how='left',left_on=['投资组合','债券'],right_on=['交易投组','成交编号/债券']).copy()
    df_bond_sell=pd.merge(df_bond_sell,df_dvlastday,how='left',left_on=['投资组合','债券'],right_on=['交易投组','成交编号/债券']).copy()
    df_bond=pd.concat([df_bond_buy,df_bond_sell],sort=False).copy()
    df_bond['PVBP']=df_bond.apply(lambda x:x['PVBP']/x['面额']*x['券面总额(万)'],axis=1)
    df_bond=df_bond[['投资组合','债券名称','交易方向','券面总额(万)','待偿期_y','到期收益率(%)','到期日','PVBP','债券类别_x']].copy()
    df_bond.insert(6,'近端日/行权日','-')
    df_bond.rename(columns={'债券名称':'交易工具','券面总额(万)':'面额/名义本金','待偿期_y':'剩余期限','到期日':'远端日/交割日/到期日','到期收益率(%)':'利率/汇率/汇差','债券类别_x':'交易类型'},inplace=True)
    df_bond['交易方向']=df_bond['交易方向'].apply(lambda x:'BUY' if x=='买入' else 'SELL' )
    df_bond['交易类型']=df_bond['交易类型'].apply(lambda x: x if x=='国债' else '政金债' if x=='政策性银行' else '其他债')
    #print(df_bond)

df_account=pd.DataFrame({'交易类型':['FR007S5Y','FR007S5Y','FR007S1Y','FR007S1Y','国债','国债','政金债','政金债','其他债','其他债'],
                         '交易方向':['BUY','SELL','BUY','SELL','BUY','SELL','BUY','SELL','BUY','SELL']})

df_displayA=pd.DataFrame()
df_displayB=pd.DataFrame()
if df_irs.empty==False or df_bond.empty==False:
    df_display=pd.concat([df_bond,df_irs])
    df_display['PVBP']=df_display.apply(lambda x:-x['PVBP'] if x['交易方向']=='SELL' and x['交易工具']!='IRS' else x['PVBP'],axis=1)
    df_displayA=df_display[df_display['投资组合'].isin(['银行间-债券IRS-黄华色-TPL(金融市场部)','银行间-债券IRS-周游力-TPL(金融市场部)','银行间-债券IRS-梁进江-TPL(金融市场部)','银行间-债券IRS-李子牛-TPL(金融市场部)'])]
    df_displayB=df_display[df_display['投资组合'].isin(['银行间-债券IRS-朱斌-TPL(金融市场部)','银行间-债券IRS-李文辉-TPL(金融市场部)','银行间-债券IRS-钟雪清-TPL(金融市场部)','银行间-债券IRS-卢奕思-TPL(金融市场部)','银行间-债券IRS-梁智航-TPL(金融市场部)','银行间-债券IRS-吴伟乐-TPL(金融市场部)'])]

    df_displayA=df_displayA.groupby(['交易类型','交易方向']).apply(lambda x:pd.Series({'名义本金':sum(x['面额/名义本金']),
                                                                                     '利率':mymean(x['利率/汇率/汇差'],x['面额/名义本金']),
                                                                                     '剩余期限':mymean(x['剩余期限'],x['面额/名义本金']),
                                                                                    'PVBP':sum(x['PVBP'])}))
    df_displayB=df_displayB.groupby(['交易类型','交易方向']).apply(lambda x:pd.Series({'名义本金':sum(x['面额/名义本金']),
                                                                                     '利率':mymean(x['利率/汇率/汇差'],x['面额/名义本金']),
                                                                                     '剩余期限':mymean(x['剩余期限'],x['面额/名义本金']),
                                                                                    'PVBP':sum(x['PVBP'])}))

    #print(df_displayA)

    #print(df_displayB)

if df_displayA.empty:
    pd_resultA=df_account
    pd_resultA['名义本金'] = np.nan
    pd_resultA['利率'] = np.nan
    pd_resultA['剩余期限'] = np.nan
    pd_resultA['PVBP'] = np.nan
else:
    pd_resultA = pd.merge(df_account, df_displayA, on=['交易类型', '交易方向'], how='left')
if df_displayB.empty:
    pd_resultB=df_account
    pd_resultB['名义本金'] = np.nan
    pd_resultB['利率'] = np.nan
    pd_resultB['剩余期限'] = np.nan
    pd_resultB['PVBP'] = np.nan
else:
    pd_resultB = pd.merge(df_account, df_displayB, on=['交易类型', '交易方向'], how='left')
print('\n')
print('-------------------------交易账户A-----------------------')
print('\n')
print(pd_resultA)

print('\n')
print('-------------------------交易账户B-----------------------')
print('\n')
print(pd_resultB)



if df_spot.empty==False:
    df_spot=df_spot[df_spot['投组'].isin(['即期-金融市场部'])]
    df_spot=df_spot[['投组','货币1金额','成交价/客户价','起息日']].copy()
    df_spot.insert(1,'交易工具','外汇即期')
    df_spot.insert(2,'交易方向',df_spot['货币1金额'].apply(lambda x: 'BUY' if x>0 else 'SELL'))
    df_spot.insert(3,'面额/名义本金',abs(df_spot['货币1金额'])/10000)
    df_spot.insert(4,'剩余期限',df_spot['起息日'].apply(lambda x: (datetime.datetime.strptime(x,'%Y-%m-%d')-day_dt).days))
    df_spot.drop('货币1金额',axis=1, inplace=True)
    df_spot.insert(6,'近端日/行权日','-')
    df_spot.rename(columns={'投组':'投资组合','成交价/客户价':'利率/汇率/汇差','起息日':'远端日/交割日/到期日'}, inplace=True)

    #print(df_spot)
if df_forward.empty==False:
    df_forward=df_forward[df_forward['投组'].isin(['远期接代-金市-结售汇','远期平盘-金市-结售汇','远期-自营-结售汇'])]
    df_forward=df_forward[['投组','货币1金额','远期全价','起息日']]
    df_forward.insert(1,'交易工具',df_forward['投组'].apply(lambda x: '远期代客' if x in ['远期接代-金市-结售汇','远期平盘-金市-结售汇'] else '远期自营' if x in ['远期-自营-结售汇'] else '' ))
    df_forward.insert(2,'交易方向',df_forward['货币1金额'].apply(lambda x: 'BUY' if x>0 else 'SELL'))
    df_forward.insert(3,'面额/名义本金',abs(df_forward['货币1金额'])/10000)
    df_forward.insert(4,'剩余期限',df_forward['起息日'].apply(lambda x: (datetime.datetime.strptime(x,'%Y-%m-%d')-day_dt).days))
    df_forward.drop('货币1金额',axis=1, inplace=True)
    df_forward.insert(6,'近端日/行权日','-')
    df_forward.rename(columns={'投组':'投资组合','成交价/客户价':'利率/汇率/汇差','起息日':'远端日/交割日/到期日','远期全价':'利率/汇率/汇差'}, inplace=True)
    #print(df_forward)
if df_swap.empty==False:
    #df_swap=df_swap[df_swap['投组'].isin(['掉期-自营-结售汇'])]
    df_swap=df_swap[['投组','近端货币1金额','价差','近端起息日','远端起息日']]
    df_swap.insert(1,'交易工具','外汇掉期')
    df_swap.insert(2,'交易方向',df_swap['近端货币1金额'].apply(lambda x: 'BUY-SELL' if x>0 else 'SELL-BUY'))
    df_swap.insert(3,'面额/名义本金',abs(df_swap['近端货币1金额']/10000))
    df_swap.insert(4,'剩余期限',df_swap['远端起息日'].apply(lambda x: (datetime.datetime.strptime(x,'%Y-%m-%d')-day_dt).days))
    df_swap.drop('近端货币1金额',axis=1, inplace=True)
    df_swap.rename(columns={'投组':'投资组合','价差':'利率/汇率/汇差','远端起息日':'远端日/交割日/到期日','近端起息日':'近端日/行权日'}, inplace=True)
    #print(df_swap)
if df_borrow.empty==False:
    #df_borrow=df_borrow[df_borrow['投组'].isin(['拆借-金市-周游力'])]
    df_borrow=df_borrow[['投组','交易方向','交易金额','利率(%)','起息日','到期还款日',]]
    df_borrow.insert(1,'交易工具','外汇拆借')
    df_borrow['交易金额']=df_borrow['交易金额']/10000
    df_borrow.insert(4,'剩余期限',df_borrow['到期还款日'].apply(lambda x: (datetime.datetime.strptime(x,'%Y-%m-%d')-day_dt).days))
    df_borrow.rename(columns={'投组':'投资组合','利率(%)':'利率/汇率/汇差','到期还款日':'远端日/交割日/到期日','起息日':'近端日/行权日','交易金额':'面额/名义本金'}, inplace=True)
    df_borrow['交易方向']=df_borrow['交易方向'].apply(lambda x:'BORROW' if x=='拆入' else 'LEND')
    #print(df_borrow)
if df_option.empty==False:
    df_option=df_option[df_option['投组'].isin(['期权-自营-结售汇','期权-自营-结售汇-周游力','期权接代-金市-外币对-周游力'])]
    df_option=df_option[['投组','交易方向','货币1金额','执行价','行权日','交割日','交易类型']]
    df_option.insert(1,'交易工具','外汇期权')
    df_option['货币1金额']=df_option['货币1金额']/10000
    df_option.insert(4,'剩余期限',df_option['交割日'].apply(lambda x: (datetime.datetime.strptime(x,'%Y-%m-%d')-day_dt).days))
    df_option['交易方向']=df_option.apply(lambda x :'BUY'+'-'+x['交易类型'] if x['交易方向']=='买入' else 'SELL'+'-'+x['交易类型'],axis=1)
    df_option.drop('交易类型',axis=1, inplace=True)
    df_option.rename(columns={'投组':'投资组合','货币1金额':'面额/名义本金','执行价':'利率/汇率/汇差','行权日':'近端日/行权日','交割日':'远端日/交割日/到期日'}, inplace=True)
    #print(df_option)
df_account_f=pd.DataFrame({'交易工具':['外汇即期','外汇即期','远期代客','远期代客','远期自营','远期自营','外汇掉期','外汇掉期','外汇拆借','外汇拆借','外汇期权','外汇期权','外汇期权','外汇期权'],
                           '交易方向':['BUY','SELL','BUY','SELL','BUY','SELL','BUY-SELL','SELL-BUY','BORROW','LEND','BUY-CALL','BUY-PUT','SELL-CALL','SELL-PUT']})
#print(df_account_f)
if df_spot.empty==False or df_forward.empty==False or df_swap.empty==False or df_borrow.empty==False or df_option.empty==False:
    df_display_f=pd.concat([df_spot,df_forward,df_swap,df_borrow,df_option],sort=False).reset_index()
    df_display_f = df_display_f.drop(labels='index', axis=1)
    df_display_f = df_display_f.groupby(['交易工具', '交易方向']).apply(lambda x: pd.Series({'名义本金': sum(x['面额/名义本金']),
                                                                                     '利率/汇率/汇差': mymean(x['利率/汇率/汇差'],
                                                                                                        x['面额/名义本金'])}))
    print(df_display_f)
else:
    df_display_f=df_account_f.copy()
    df_display_f['名义本金'] = np.nan
    df_display_f['利率/汇率/汇差'] = np.nan






pd_result_f=pd.merge(df_account_f,df_display_f,on=['交易工具','交易方向'],how='left')
print('\n')
print('-------------------------外汇交易-----------------------')
print('\n')
print(pd_result_f)






df_trade=pd.concat([df_irs,df_bond],sort=False).iloc[:,:8].copy()
df_trade=pd.concat([df_trade,df_spot,df_forward,df_swap,df_borrow,df_option],sort=False).reset_index()
df_trade=df_trade.drop(labels='index',axis=1)
print('\n')
print('------------------------------------------------------------总交易记录----------------------------------------------------------')
print('\n')
print(df_trade)
book=load_workbook(outaddress)
writer=pd.ExcelWriter(outaddress,engine='openpyxl')
writer.book=book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
ws=book['交易记录']
for row in ws['N3:V100']:
    for cell in row:
        cell.value=''

df_trade.to_excel(writer,sheet_name='交易记录',startcol=13,startrow=1)
pd_resultA2=pd_resultA.drop(['交易类型','交易方向'],axis=1)
pd_resultA2.to_excel(writer,sheet_name='交易记录',startcol=3,startrow=1,index=False)
pd_resultB2=pd_resultB.drop(['交易类型','交易方向'],axis=1)
pd_resultB2.to_excel(writer,sheet_name='交易记录',startcol=3,startrow=15,index=False)
pd_result_f2=pd_result_f.drop(['交易工具','交易方向'],axis=1)
pd_result_f2.to_excel(writer,sheet_name='交易记录',startcol=10,startrow=1,index=False)
ws=book['DV01（底稿）']
for i in range(df_dvtoday.shape[0],df_dvtoday.shape[0]+50):
    for j in range(1,32):
        ws.cell(row=i,column=j).value=''
df_dvtoday.to_excel(writer,sheet_name='DV01（底稿）',startrow=2,index=False)
ws=book['年损益（底稿）']
for i in range(df_year.shape[0],df_year.shape[0]+50):
    for j in range(1,57):
        ws.cell(row=i,column=j).value=''
df_year.to_excel(writer,sheet_name='年损益（底稿）',startrow=2,index=False)
ws=book['季损益（底稿）']
for i in range(df_quarter.shape[0],df_quarter.shape[0]+50):
    for j in range(1,57):
        ws.cell(row=i,column=j).value=''
df_quarter.to_excel(writer,sheet_name='季损益（底稿）',startrow=2,index=False)
ws=book['外汇DV01（底稿）']
for i in range(1,25):
    for j in range(1,30):
        ws.cell(row=i,column=j).value=''

df_fxpositon.to_excel(writer,sheet_name='外汇DV01（底稿）',index=False)

ws=book['外汇损益（底稿）']
for i in range(1,30):
    for j in range(1,30):
        ws.cell(row=i,column=j).value=''
df_fxgain.to_excel(writer,sheet_name='外汇损益（底稿）',index=False)
ws=book['日报']
ws.cell(row=1,column=2).value=day_dt
book.save(outaddress)
writer.save()
